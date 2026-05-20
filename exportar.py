import sys, json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.loads(sys.stdin.read())
pedidos = data.get("pedidos", [])
apontamentos = data.get("apontamentos", [])

wb = Workbook()

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def today_str():
    return datetime.date.today().strftime("%d/%m/%Y")

def excel_date(d):
    if not d: return ""
    try: return datetime.datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except: return d

def calc_status(p):
    if p.get("qtdProduzida", 0) >= p.get("qtdPedida", 0): return "Concluído"
    try:
        prazo = datetime.datetime.strptime(p["prazoEntrega"][:10], "%Y-%m-%d").date()
        diff = (prazo - datetime.date.today()).days
        if diff < 0: return "Atrasado"
        if diff <= 2: return "Em risco"
        return "No prazo"
    except: return "—"

STATUS_COLORS = {
    "Concluído": ("1A7431", "D4EDDA"),
    "No prazo":  ("0056A2", "D0E8FF"),
    "Em risco":  ("856404", "FFF3CD"),
    "Atrasado":  ("842029", "F8D7DA"),
}

# ABA PEDIDOS
ws1 = wb.active
ws1.title = "Pedidos"
ws1.sheet_view.showGridLines = False
ws1.merge_cells("A1:L1")
ws1["A1"] = "CONTROLE DE PEDIDOS — PRODUÇÃO"
ws1["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
ws1["A1"].fill = fill("1C3557")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:L2")
ws1["A2"] = f"Gerado em: {today_str()}   |   Total de ordens: {len(pedidos)}"
ws1["A2"].font = Font(color="FFFFFF", name="Arial", size=9, italic=True)
ws1["A2"].fill = fill("2A4A72")
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

headers = ["Nº Pedido","Produto","Cliente","Qtd Pedida","Unid.","Data Pedido","Prazo Entrega","Dias Restantes","Qtd Produzida","% Concluído","Saldo","Status"]
col_w   = [12, 24, 18, 12, 8, 14, 14, 14, 14, 13, 10, 14]
ws1.row_dimensions[3].height = 20
for col, (h, w) in enumerate(zip(headers, col_w), 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill = fill("2C5F8A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    ws1.column_dimensions[get_column_letter(col)].width = w

for r, p in enumerate(pedidos, 4):
    ws1.row_dimensions[r].height = 18
    status = calc_status(p)
    try:
        prazo = datetime.datetime.strptime(p["prazoEntrega"][:10], "%Y-%m-%d").date()
        dias = (prazo - datetime.date.today()).days
    except: dias = 0
    pct = round(p.get("qtdProduzida",0)/p["qtdPedida"]*100, 1) if p["qtdPedida"] else 0
    saldo = p["qtdPedida"] - p.get("qtdProduzida", 0)
    sc, sf = STATUS_COLORS.get(status, ("000000","FFFFFF"))
    row_vals = [p["id"],p["produto"],p["cliente"],p["qtdPedida"],p.get("unid","und"),
                excel_date(p.get("dataPedido","")),excel_date(p.get("prazoEntrega","")),
                dias,p.get("qtdProduzida",0),f"{pct}%",saldo,status]
    for col, val in enumerate(row_vals, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=10, bold=(col==12), color=sc if col==12 else "000000")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col == 12: c.fill = fill(sf)
        elif r % 2 == 0: c.fill = fill("F4F7FB")

# ABA APONTAMENTOS
ws2 = wb.create_sheet("Apontamento Diário")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:I1")
ws2["A1"] = "APONTAMENTO DIÁRIO DE PRODUÇÃO"
ws2["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
ws2["A1"].fill = fill("1A5C38")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

ap_headers = ["Data/Hora","Nº OP","Produto","Cliente","Turno","Operador","Qtd Produzida","Unid.","Observações"]
ap_widths   = [18, 12, 24, 18, 10, 16, 14, 8, 36]
ws2.row_dimensions[2].height = 20
for col, (h, w) in enumerate(zip(ap_headers, ap_widths), 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill = fill("2E7D52")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    ws2.column_dimensions[get_column_letter(col)].width = w

for r, ap in enumerate(apontamentos, 3):
    ws2.row_dimensions[r].height = 18
    try: dt = datetime.datetime.fromisoformat(ap["timestamp"]).strftime("%d/%m/%Y %H:%M")
    except: dt = ap.get("data", "")
    row_vals = [dt,ap.get("pedidoId",""),ap.get("produto",""),ap.get("cliente",""),
                ap.get("turno",""),ap.get("operador",""),ap.get("qtd",0),ap.get("unid",""),ap.get("obs","")]
    for col, val in enumerate(row_vals, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center" if col in (7,8) else "left", vertical="center")
        if r % 2 == 0: c.fill = fill("EEF7F2")

# ABA DASHBOARD
ws3 = wb.create_sheet("Dashboard")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:I1")
ws3["A1"] = "DASHBOARD — RESUMO EXECUTIVO"
ws3["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=14)
ws3["A1"].fill = fill("1C3557")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

ws3.merge_cells("A2:I2")
ws3["A2"] = f"Exportado em: {today_str()}"
ws3["A2"].font = Font(color="FFFFFF", name="Arial", size=9, italic=True)
ws3["A2"].fill = fill("2A4A72")
ws3["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 18

total = len(pedidos)
conc  = sum(1 for p in pedidos if calc_status(p)=="Concluído")
atr   = sum(1 for p in pedidos if calc_status(p)=="Atrasado")
risco = sum(1 for p in pedidos if calc_status(p)=="Em risco")
prazoN= sum(1 for p in pedidos if calc_status(p)=="No prazo")
tot_prod = sum(p.get("qtdProduzida",0) for p in pedidos)
tot_ped  = sum(p["qtdPedida"] for p in pedidos)
geral = round(tot_prod/tot_ped*100,1) if tot_ped else 0

kpis = [("Total OPs",total,"2C5F8A"),("Concluídos",conc,"1A7431"),("No Prazo",prazoN,"0056A2"),("Em Risco",risco,"856404"),("Atrasados",atr,"842029")]
for i,(lbl,val,cor) in enumerate(kpis):
    col = i*2+1
    ws3.column_dimensions[get_column_letter(col)].width = 14
    ws3.column_dimensions[get_column_letter(col+1)].width = 2
    ws3.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1)
    ws3.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
    lc = ws3.cell(row=4,column=col,value=lbl)
    lc.font = Font(name="Arial",size=9,color="888888")
    lc.alignment = Alignment(horizontal="center")
    vc = ws3.cell(row=5,column=col,value=val)
    vc.font = Font(bold=True,name="Arial",size=22,color="FFFFFF")
    vc.fill = fill(cor)
    vc.alignment = Alignment(horizontal="center",vertical="center")
    ws3.row_dimensions[5].height = 40

ws3.merge_cells("A7:J7")
pg = ws3["A7"]
pg.value = f"Progresso geral: {geral}%  ({tot_prod:,} / {tot_ped:,} unidades produzidas)"
pg.font = Font(bold=True,name="Arial",size=11,color="1C3557")
pg.alignment = Alignment(horizontal="center",vertical="center")
pg.fill = fill("EEF4FF")
ws3.row_dimensions[7].height = 24

dash_h = ["Nº OP","Produto","Cliente","Qtd Total","Produzido","Saldo","% Concluído","Prazo","Status"]
dash_w = [10,22,16,10,10,10,12,14,14]
ws3.row_dimensions[9].height = 20
for col,(h,w) in enumerate(zip(dash_h,dash_w),1):
    c = ws3.cell(row=9,column=col,value=h)
    c.font = Font(bold=True,color="FFFFFF",name="Arial",size=10)
    c.fill = fill("2C5F8A")
    c.alignment = Alignment(horizontal="center",vertical="center")
    c.border = thin_border()
    ws3.column_dimensions[get_column_letter(col)].width = w

for r,p in enumerate(pedidos,10):
    ws3.row_dimensions[r].height = 18
    status = calc_status(p)
    sc,sf = STATUS_COLORS.get(status,("000000","FFFFFF"))
    pct_v = round(p.get("qtdProduzida",0)/p["qtdPedida"]*100,1) if p["qtdPedida"] else 0
    saldo = p["qtdPedida"]-p.get("qtdProduzida",0)
    row_vals=[p["id"],p["produto"],p["cliente"],p["qtdPedida"],p.get("qtdProduzida",0),saldo,f"{pct_v}%",excel_date(p.get("prazoEntrega","")),status]
    for col,val in enumerate(row_vals,1):
        c = ws3.cell(row=r,column=col,value=val)
        c.font = Font(bold=(col==9),name="Arial",size=10,color=sc if col==9 else "000000")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center",vertical="center")
        if col==9: c.fill=fill(sf)
        elif r%2==0: c.fill=fill("F4F7FB")

out = "relatorio_producao.xlsx"
wb.save(out)
print(f"Arquivo salvo: {out}")
